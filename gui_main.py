# -*- coding: utf-8 -*-

"""
加班数据提取和处理 GUI 应用程序

该应用程序提供图形界面用于从Excel文件中提取指定研究室的加班数据，
并生成格式化的结果文件。

功能包括：
1. 选择源数据文件（.xlsx或.xls）
2. 拖放文件支持
3. 选择结果文件保存位置
4. 选择要提取的研究室
5. 执行数据提取和处理
"""

import sys
import os
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QLabel, QPushButton, QFileDialog, QComboBox, QTextEdit, 
    QMessageBox, QGroupBox, QGridLayout, QFrame, QCheckBox
)
from PyQt5.QtCore import Qt, QMimeData
from PyQt5.QtGui import QDragEnterEvent, QDropEvent, QFont


class DragDropLabel(QLabel):
    """支持拖放文件的标签"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setAlignment(Qt.AlignCenter)
        self.setText("将Excel文件拖放到这里\n或点击下方按钮选择文件")
        self.setStyleSheet("""
            QLabel {
                border: 2px dashed #aaa;
                border-radius: 10px;
                padding: 20px;
                margin: 10px;
                background-color: #f0f0f0;
            }
        """)
    
    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            # 先接受拖拽动作，具体的文件类型检查在dropEvent中进行
            event.acceptProposedAction()
        else:
            event.ignore()
    
    def dropEvent(self, event: QDropEvent):
        if event.mimeData().hasUrls():
            file_path = event.mimeData().urls()[0].toLocalFile()
            if file_path.lower().endswith(('.xlsx', '.xls')):
                # Find the main window to call set_source_file
                main_window = self.window()
                if hasattr(main_window, 'set_source_file'):
                    main_window.set_source_file(file_path)
            else:
                # Find the main window for the message box
                main_window = self.window()
                QMessageBox.warning(main_window, "文件类型错误", "请选择Excel文件（.xlsx或.xls）")


class OvertimeExtractorApp(QMainWindow):
    """加班数据提取应用程序主窗口"""
    
    def __init__(self):
        super().__init__()
        self.source_file = ""
        self.output_dir = ""
        self.selected_labs = []  # 改为列表以支持多选
        self.research_labs = ['智能通信', '数据算法', '新型能源', '新型材料']
        self.init_ui()
    
    def init_ui(self):
        """初始化用户界面"""
        self.setWindowTitle("加班数据提取工具")
        self.setGeometry(100, 100, 600, 500)
        
        # 创建中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        
        # 标题
        title_label = QLabel("加班数据提取工具")
        title_label.setAlignment(Qt.AlignCenter)
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title_label.setFont(title_font)
        main_layout.addWidget(title_label)
        
        # 文件选择区域
        file_group = QGroupBox("文件选择")
        file_layout = QVBoxLayout(file_group)
        
        # 拖放区域
        self.drag_drop_label = DragDropLabel(self)
        file_layout.addWidget(self.drag_drop_label)
        
        # 文件选择按钮
        file_button_layout = QHBoxLayout()
        self.select_file_btn = QPushButton("选择源文件")
        self.select_file_btn.clicked.connect(self.select_source_file)
        self.select_output_btn = QPushButton("选择保存位置")
        self.select_output_btn.clicked.connect(self.select_output_dir)
        file_button_layout.addWidget(self.select_file_btn)
        file_button_layout.addWidget(self.select_output_btn)
        file_layout.addLayout(file_button_layout)
        
        # 文件路径显示
        self.source_file_label = QLabel("源文件: 未选择")
        self.output_dir_label = QLabel("保存位置: 当前目录")
        file_layout.addWidget(self.source_file_label)
        file_layout.addWidget(self.output_dir_label)
        
        main_layout.addWidget(file_group)
        
        # 研究室选择区域
        lab_group = QGroupBox("研究室选择")
        lab_layout = QVBoxLayout(lab_group)
        
        # 创建复选框用于选择研究室
        self.lab_checkboxes = []
        for lab in self.research_labs:
            checkbox = QCheckBox(lab)
            self.lab_checkboxes.append(checkbox)
            lab_layout.addWidget(checkbox)
        
        main_layout.addWidget(lab_group)
        
        # 操作按钮
        button_layout = QHBoxLayout()
        self.process_btn = QPushButton("开始处理")
        self.process_btn.clicked.connect(self.process_data)
        self.exit_btn = QPushButton("退出")
        self.exit_btn.clicked.connect(self.close)
        button_layout.addWidget(self.process_btn)
        button_layout.addWidget(self.exit_btn)
        main_layout.addLayout(button_layout)
        
        # 日志输出区域
        log_group = QGroupBox("处理日志")
        log_layout = QVBoxLayout(log_group)
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        log_layout.addWidget(self.log_text)
        
        main_layout.addWidget(log_group)
        
        # 状态栏
        self.statusBar().showMessage("就绪")
    
    def set_source_file(self, file_path):
        """设置源文件路径"""
        self.source_file = file_path
        self.source_file_label.setText(f"源文件: {os.path.basename(file_path)}")
        self.statusBar().showMessage(f"已选择源文件: {file_path}")
        self.log_message(f"已选择源文件: {file_path}")
    
    def select_source_file(self):
        """选择源文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            "选择源文件", 
            "", 
            "Excel文件 (*.xlsx *.xls)"
        )
        if file_path:
            self.set_source_file(file_path)
    
    def select_output_dir(self):
        """选择输出目录"""
        dir_path = QFileDialog.getExistingDirectory(self, "选择保存位置")
        if dir_path:
            self.output_dir = dir_path
            self.output_dir_label.setText(f"保存位置: {dir_path}")
            self.statusBar().showMessage(f"已选择保存位置: {dir_path}")
            self.log_message(f"已选择保存位置: {dir_path}")
    
    def log_message(self, message):
        """添加日志消息"""
        self.log_text.append(message)
        self.log_text.verticalScrollBar().setValue(
            self.log_text.verticalScrollBar().maximum()
        )
        QApplication.processEvents()  # 确保界面及时更新
    
    def format_overtime_records(self, record):
        """格式化加班记录"""
        if pd.isna(record) or record == '[]':
            return ''
        
        # 使用正则表达式提取完整的日期
        dates = re.findall(r'\d{4}-(\d{2})-(\d{2})', record)
        
        # 格式化日期，去除前导零
        formatted_dates = [f'{int(month)}月{int(day)}日加班' for month, day in dates]
        
        # 按日期排序
        formatted_dates.sort(key=lambda x: int(x.split('月')[1].split('日')[0]))
        
        return '\n'.join(formatted_dates)
    
    def process_data(self):
        """处理数据"""
        # 检查必要条件
        if not self.source_file:
            QMessageBox.warning(self, "警告", "请选择源文件")
            return
        
        if not os.path.exists(self.source_file):
            QMessageBox.warning(self, "警告", "源文件不存在")
            return
        
        # 获取选择的研究室
        self.selected_labs = [cb.text() for cb in self.lab_checkboxes if cb.isChecked()]
        if not self.selected_labs:
            QMessageBox.warning(self, "警告", "请至少选择一个研究室")
            return
            
        self.log_message(f"开始处理 {', '.join(self.selected_labs)} 的数据...")
        
        try:
            # 读取源数据文件
            self.log_message("正在读取源数据文件...")
            df = pd.read_excel(self.source_file)
            
            # 根据选择的研究室过滤数据
            filtered_df = df[df['研究室'].isin(self.selected_labs)]
            
            # 数据处理和整理
            self.log_message("正在处理数据...")
            result_df = filtered_df[['姓名', '周末加班记录', '周末加班奖励总额', '研究室']]
            
            # 去除"周末加班记录"为空的行
            result_df = result_df[result_df['周末加班记录'].astype(str) != '[]']
            
            # 整理"奖惩明细"栏内容
            result_df['奖惩明细'] = result_df['周末加班记录'].apply(self.format_overtime_records)
            
            # 添加开发室列
            result_df['开发室'] = result_df['研究室'].apply(lambda x: f'{x}研究室')
            
            # 按照指定的研究室顺序排序
            # 研究室顺序：智能通信>数据算法>新型能源>新型材料
            research_lab_order = ['智能通信', '数据算法', '新型能源', '新型材料']
            
            # 创建一个映射字典来定义研究室的排序顺序
            lab_order_dict = {lab: i for i, lab in enumerate(research_lab_order)}
            
            # 添加一个临时列用于排序
            result_df['研究室排序'] = result_df['研究室'].map(lab_order_dict)
            
            # 添加一个索引列来保持源文件中的顺序
            result_df['源文件索引'] = result_df.index
            
            # 按研究室顺序和源文件中的顺序排序
            result_df = result_df.sort_values(['研究室排序', '源文件索引'])
            
            # 删除临时排序列
            result_df.drop(['研究室排序', '源文件索引'], axis=1, inplace=True)
            
            # 重命名列
            result_df.rename(columns={'周末加班奖励总额': '建议奖励金额', '姓名': '涉及人员'}, inplace=True)
            
            # 添加序号列（在排序后添加，确保序号连续）
            result_df.insert(0, '序号', range(1, len(result_df) + 1))
            
            # 添加奖惩项点列
            result_df.insert(1, '奖惩项点', '节假日交通奖励')
            
            # 添加建议处罚金额列
            result_df['建议处罚金额'] = ''
            
            # 删除研究室列（在排序之后）
            result_df.drop('研究室', axis=1, inplace=True)
            
            # 调整列顺序
            result_df = result_df[['序号', '奖惩项点', '奖惩明细', '建议奖励金额', '建议处罚金额', '涉及人员', '开发室']]
            
            # 创建并格式化 Excel 文件
            self.log_message("正在创建结果文件...")
            wb = Workbook()
            ws = wb.active
            
            # 将 DataFrame 写入工作表
            for r_idx, row in enumerate(dataframe_to_rows(result_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            # 设置字体和背景色
            header_font = Font(name='微软雅黑', size=11, bold=True)
            header_fill = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            body_font = Font(name='宋体', size=11)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = body_font
            
            # 设置对齐方式和自动换行
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 添加单元格框线
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell.border = thin_border
            
            # 合并单元格
            if len(result_df) > 1:
                # 合并序号列
                ws.merge_cells(start_row=2, start_column=1, end_row=len(result_df) + 1, end_column=1)
                ws.cell(row=2, column=1, value=1)
                ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row=2, column=1).border = thin_border
                
                # 合并奖惩项点列
                ws.merge_cells(start_row=2, start_column=2, end_row=len(result_df) + 1, end_column=2)
                ws.cell(row=2, column=2, value="节假日交通奖励")
                ws.cell(row=2, column=2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row=2, column=2).border = thin_border
                
                # 合并相同开发室的单元格
                # 获取开发室列的数据
                dev_rooms = result_df['开发室'].tolist()
                start_row = 2  # 数据从第2行开始
                
                # 遍历开发室列表，找到连续相同的开发室并合并
                i = 0
                while i < len(dev_rooms):
                    current_room = dev_rooms[i]
                    start_index = i
                    
                    # 找到连续相同的开发室
                    while i < len(dev_rooms) and dev_rooms[i] == current_room:
                        i += 1
                    
                    # 如果有多个连续相同的开发室，则合并
                    if i - start_index > 1:
                        # 合并开发室列的单元格 (第7列)
                        ws.merge_cells(
                            start_row=start_row + start_index,
                            start_column=7,
                            end_row=start_row + i - 1,
                            end_column=7
                        )
                        
                        # 设置合并后单元格的值和格式
                        cell = ws.cell(row=start_row + start_index, column=7, value=current_room)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = thin_border
                    else:
                        # 单个开发室，只设置格式
                        cell = ws.cell(row=start_row + start_index, column=7, value=current_room)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = thin_border
            
            # 设置固定的列宽
            ws.column_dimensions['A'].width = 10   # 序号列
            ws.column_dimensions['B'].width = 15   # 奖惩项点列
            ws.column_dimensions['C'].width = 25   # 奖惩明细列
            ws.column_dimensions['D'].width = 15   # 建议奖励金额列
            ws.column_dimensions['E'].width = 15   # 建议处罚金额列
            ws.column_dimensions['F'].width = 15   # 涉及人员列
            ws.column_dimensions['G'].width = 20   # 开发室列
            
            # 自动调整行高（基于内容）
            for row in ws.iter_rows():
                max_height = 15  # 默认行高
                for cell in row:
                    if cell.value:
                        # 计算行高，假设每行15磅，每换行增加15磅
                        lines = str(cell.value).count('\n') + 1
                        cell_height = lines * 15
                        if cell_height > max_height:
                            max_height = cell_height
                ws.row_dimensions[row[0].row].height = max_height
            
            # 保存文件
            output_path = os.path.join(
                self.output_dir if self.output_dir else os.path.dirname(self.source_file),
                "result.xlsx"
            )
            
            wb.save(output_path)
            self.log_message(f"数据处理完成，结果已保存到: {output_path}")
            QMessageBox.information(self, "完成", f"数据提取完成！\n结果已保存到: {output_path}")
            self.statusBar().showMessage("处理完成")
            
        except Exception as e:
            error_msg = f"处理过程中发生错误: {str(e)}"
            self.log_message(error_msg)
            QMessageBox.critical(self, "错误", error_msg)
            self.statusBar().showMessage("处理失败")


def main():
    app = QApplication(sys.argv)
    window = OvertimeExtractorApp()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()