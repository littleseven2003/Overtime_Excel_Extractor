import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# 读取source.xlsx文件
df = pd.read_excel('source.xlsx')

# 定义研究室选项
research_labs = ['智能通信', '数据算法', '新型能源', '新型材料']

# 提示用户选择研究室
print("请选择研究室：")
for i, lab in enumerate(research_labs, 1):
    print(f"{i}. {lab}")

choice = int(input("请输入选项编号: "))
selected_lab = research_labs[choice - 1]

# 根据选择的研究室过滤数据
filtered_df = df[df['研究室'] == selected_lab]

# 提取姓名、周末加班记录、周末加班奖励总额
# 假设列名分别为 '姓名', '周末加班记录', '周末加班奖励总额'
result_df = filtered_df[['姓名', '周末加班记录', '周末加班奖励总额']]

# 去除“周末加班记录”为空（即内容为[]）的行
result_df = result_df[result_df['周末加班记录'].astype(str) != '[]']

# 整理“奖惩明细”栏内容
import re
def format_overtime_records(record):
    if pd.isna(record) or record == '[]':
        return ''
    # 使用正则表达式提取完整的日期
    dates = re.findall(r'\d{4}-(\d{2})-(\d{2})', record)
    # 格式化日期，去除前导零
    formatted_dates = [f'{int(month)}月{int(day)}日加班' for month, day in dates]
    # 按日期排序
    formatted_dates.sort(key=lambda x: int(x.split('月')[1].split('日')[0]))
    return '\n'.join(formatted_dates)

result_df['奖惩明细'] = result_df['周末加班记录'].apply(format_overtime_records)

# 调整列顺序，将姓名列移到最后
result_df = result_df[['奖惩明细', '周末加班奖励总额', '姓名']]

# 添加序号列
result_df.insert(0, '序号', [1] + [''] * (len(result_df) - 1))

# 添加奖惩项点列
result_df.insert(1, '奖惩项点', ['节假日交通奖励'] + [''] * (len(result_df) - 1))

# 重命名列
result_df.rename(columns={'周末加班奖励总额': '建议奖励金额', '姓名': '涉及人员'}, inplace=True)

# 添加建议处罚金额列
result_df['建议处罚金额'] = ''

# 添加开发室列
result_df['开发室'] = [f'{selected_lab}研究室'] + [''] * (len(result_df) - 1)

# 调整列顺序
result_df = result_df[['序号', '奖惩项点', '奖惩明细', '建议奖励金额', '建议处罚金额', '涉及人员', '开发室']]

# 将结果写入result.xlsx文件
# 使用 openpyxl 创建工作簿和工作表
wb = Workbook()
ws = wb.active

# 将 DataFrame 写入工作表
for r_idx, row in enumerate(dataframe_to_rows(result_df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# 设置标题行格式
from openpyxl.styles import Font, PatternFill
header_font = Font(name='微软雅黑', size=11, bold=True)
header_fill = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill

# 设置其他行格式
body_font = Font(name='宋体', size=11)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.font = body_font

# 设置所有单元格的对齐方式和自动换行
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 给有内容的单元格添加框线
from openpyxl.styles import Border, Side
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for row in ws.iter_rows():
    for cell in row:
        # 检查单元格是否有内容（包括空字符串）
        if cell.value is not None:
            cell.border = thin_border

# 合并序号列的单元格 (A2 到 A(最后一行))
if len(result_df) > 1:
    ws.merge_cells(start_row=2, start_column=1, end_row=len(result_df) + 1, end_column=1)
    # 设置合并后单元格的值为1
    ws.cell(row=2, column=1, value=1)
    # 设置单元格内容居中
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # 添加框线
    ws.cell(row=2, column=1).border = thin_border

# 合并奖惩项点列的单元格 (B2 到 B(最后一行))
if len(result_df) > 1:
    ws.merge_cells(start_row=2, start_column=2, end_row=len(result_df) + 1, end_column=2)
    # 设置合并后单元格的值为"节假日交通奖励"
    ws.cell(row=2, column=2, value="节假日交通奖励")
    # 设置单元格内容居中
    ws.cell(row=2, column=2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # 添加框线
    ws.cell(row=2, column=2).border = thin_border

# 合并开发室列的单元格 (G2 到 G(最后一行))
if len(result_df) > 1:
    ws.merge_cells(start_row=2, start_column=7, end_row=len(result_df) + 1, end_column=7)
    # 设置合并后单元格的值为研究室名称
    ws.cell(row=2, column=7, value=f'{selected_lab}研究室')
    # 设置单元格内容居中
    ws.cell(row=2, column=7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # 添加框线
    ws.cell(row=2, column=7).border = thin_border

# 自动调整列宽
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    
    # 调整特定列的宽度
    if column_letter == 'C':  # 奖惩明细列
        ws.column_dimensions[column_letter].width = max(adjusted_width * 0.8, 10)  # 减少宽度
    elif column_letter in ['B', 'F', 'D', 'E']:  # 奖惩项点、涉及人员、奖励金额、惩罚金额列
        ws.column_dimensions[column_letter].width = max(adjusted_width * 1.2, 15)  # 增加宽度
    else:
        ws.column_dimensions[column_letter].width = adjusted_width

# 自动调整行高
for row in ws.iter_rows():
    max_height = 15  # 默认行高
    for cell in row:
        if cell.value:
            # 计算行高，假设每行15磅，每换行增加15磅
            lines = str(cell.value).count('\n') + 1
            cell_height = lines * 15
            if cell_height > max_height:
                max_height = cell_height
    ws.row_dimensions[row[0].row].height = max_height

# 保存 Excel 文件
wb.save('result.xlsx')

print(f"数据提取完成并写入result.xlsx文件。已选择研究室: {selected_lab}")