#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
加班数据提取工具 - 核心业务逻辑模块

该模块负责处理加班数据的核心业务逻辑，包括：
1. 读取源Excel文件
2. 根据选择的研究室过滤数据
3. 格式化加班记录
4. 生成格式化的结果Excel文件
"""

import pandas as pd
import re
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class OvertimeDataProcessor:
    """加班数据处理器类
    
    负责处理加班数据的核心业务逻辑，包括数据读取、过滤、格式化和输出。
    
    Attributes:
        source_file (str): 源Excel文件路径
        output_dir (str): 输出目录路径
        selected_labs (list): 选中的研究室列表
    """
    
    def __init__(self, source_file, output_dir="", selected_labs=None):
        """初始化加班数据处理器
        
        Args:
            source_file (str): 源Excel文件路径
            output_dir (str, optional): 输出目录路径. Defaults to "".
            selected_labs (list, optional): 选中的研究室列表. Defaults to None.
        """
        self.source_file = source_file
        self.output_dir = output_dir
        self.selected_labs = selected_labs or []
    
    def format_overtime_records(self, record):
        """格式化加班记录
        
        将原始加班记录格式化为指定格式的字符串，例如：
        "[2023-05-01, 2023-05-07]" -> "5月1日加班\n5月7日加班"
        
        Args:
            record (str): 原始加班记录字符串
            
        Returns:
            str: 格式化后的加班记录字符串
        """
        # 处理空值或空列表情况
        if pd.isna(record) or record == '[]':
            return ''
        
        # 使用正则表达式提取完整的日期 (年-月-日)
        dates = re.findall(r'\d{4}-(\d{2})-(\d{2})', record)
        
        # 格式化日期，去除前导零 (例如: 05 -> 5)
        formatted_dates = [f'{int(month)}月{int(day)}日加班' for month, day in dates]
        
        # 按日期排序，确保输出结果按日期顺序排列
        formatted_dates.sort(key=lambda x: int(x.split('月')[1].split('日')[0]))
        
        # 用换行符连接所有日期
        return '\n'.join(formatted_dates)
    
    def process_data(self, log_callback=None):
        """处理加班数据并生成结果文件
        
        主要处理流程：
        1. 读取源数据文件
        2. 根据选择的研究室过滤数据
        3. 格式化加班记录
        4. 重新排序和组织数据
        5. 生成格式化的Excel结果文件
        
        Args:
            log_callback (callable, optional): 日志回调函数，用于输出处理过程信息. Defaults to None.
            
        Returns:
            str: 生成的结果文件路径
            
        Raises:
            Exception: 处理过程中发生错误时抛出异常
        """
        # 记录开始处理的日志信息
        if log_callback:
            log_callback(f"开始处理 {', '.join(self.selected_labs)} 的数据...")
        
        try:
            # 1. 读取源数据文件
            if log_callback:
                log_callback("正在读取源数据文件...")
            df = pd.read_excel(self.source_file)
            
            # 2. 根据选择的研究室过滤数据
            filtered_df = df[df['研究室'].isin(self.selected_labs)]
            
            # 3. 数据处理和整理
            if log_callback:
                log_callback("正在处理数据...")
            result_df = filtered_df[['姓名', '周末加班记录', '周末加班奖励总额', '研究室']]
            
            # 去除"周末加班记录"为空的行（即内容为[]的行）
            result_df = result_df[result_df['周末加班记录'].astype(str) != '[]']
            
            # 整理"奖惩明细"栏内容，使用自定义格式化函数
            result_df['奖惩明细'] = result_df['周末加班记录'].apply(self.format_overtime_records)
            
            # 添加开发室列，格式为"研究室名称+研究室"
            result_df['开发室'] = result_df['研究室'].apply(lambda x: f'{x}研究室')
            
            # 4. 按照指定的研究室顺序排序
            # 研究室顺序：智能通信>数据算法>新型能源>新型材料>智能装备
            research_lab_order = ['智能通信', '数据算法', '新型能源', '新型材料', '智能装备']
            
            # 创建一个映射字典来定义研究室的排序顺序
            lab_order_dict = {lab: i for i, lab in enumerate(research_lab_order)}
            
            # 添加一个临时列用于排序
            result_df['研究室排序'] = result_df['研究室'].map(lab_order_dict)
            
            # 添加一个索引列来保持源文件中的顺序
            result_df['源文件索引'] = result_df.index
            
            # 按研究室顺序和源文件中的顺序排序
            result_df = result_df.sort_values(['研究室排序', '源文件索引'])
            
            # 删除临时排序列
            result_df.drop(['研究室排序', '源文件索引'], axis=1, inplace=True)
            
            # 重命名列，使其更符合业务需求
            result_df.rename(columns={
                '周末加班奖励总额': '建议奖励金额', 
                '姓名': '涉及人员'
            }, inplace=True)
            
            # 添加序号列（在排序后添加，确保序号连续）
            result_df.insert(0, '序号', range(1, len(result_df) + 1))
            
            # 添加奖惩项点列
            result_df.insert(1, '奖惩项点', '节假日交通奖励')
            
            # 添加建议处罚金额列（初始为空）
            result_df['建议处罚金额'] = ''
            
            # 删除研究室列（在排序之后，已不再需要）
            result_df.drop('研究室', axis=1, inplace=True)
            
            # 调整最终列顺序
            result_df = result_df[[
                '序号', '奖惩项点', '奖惩明细', '建议奖励金额', 
                '建议处罚金额', '涉及人员', '开发室'
            ]]
            
            # 5. 创建并格式化 Excel 文件
            if log_callback:
                log_callback("正在创建结果文件...")
            wb = Workbook()
            ws = wb.active
            
            # 将 DataFrame 写入工作表
            for r_idx, row in enumerate(dataframe_to_rows(result_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            # 设置字体和背景色
            # 标题行：微软雅黑、11号、粗体、浅蓝色背景
            header_font = Font(name='微软雅黑', size=11, bold=True)
            header_fill = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
            
            for cell in ws[1]:  # 第一行是标题行
                cell.font = header_font
                cell.fill = header_fill
            
            # 数据行：宋体、11号
            body_font = Font(name='宋体', size=11)
            for row in ws.iter_rows(min_row=2):  # 从第二行开始是数据行
                for cell in row:
                    cell.font = body_font
            
            # 设置对齐方式和自动换行
            # 所有单元格：水平居中、垂直居中、自动换行
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 添加单元格框线
            # 所有有内容的单元格添加细线边框
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:  # 只给有内容的单元格添加边框
                        cell.border = thin_border
            
            # 合并单元格
            if len(result_df) > 1:  # 只有在有多行数据时才需要合并
                # 合并序号列（第1列）
                ws.merge_cells(
                    start_row=2, 
                    start_column=1, 
                    end_row=len(result_df) + 1, 
                    end_column=1
                )
                ws.cell(row=2, column=1, value=1)  # 合并后的单元格值为1
                ws.cell(row=2, column=1).alignment = Alignment(
                    horizontal='center', 
                    vertical='center', 
                    wrap_text=True
                )
                ws.cell(row=2, column=1).border = thin_border
                
                # 合并奖惩项点列（第2列）
                ws.merge_cells(
                    start_row=2, 
                    start_column=2, 
                    end_row=len(result_df) + 1, 
                    end_column=2
                )
                ws.cell(row=2, column=2, value="节假日交通奖励")  # 合并后的单元格值
                ws.cell(row=2, column=2).alignment = Alignment(
                    horizontal='center', 
                    vertical='center', 
                    wrap_text=True
                )
                ws.cell(row=2, column=2).border = thin_border
                
                # 合并相同开发室的单元格（第7列）
                # 获取开发室列的数据
                dev_rooms = result_df['开发室'].tolist()
                start_row = 2  # 数据从第2行开始
                
                # 遍历开发室列表，找到连续相同的开发室并合并
                i = 0
                while i < len(dev_rooms):
                    current_room = dev_rooms[i]
                    start_index = i
                    
                    # 找到连续相同的开发室
                    while i < len(dev_rooms) and dev_rooms[i] == current_room:
                        i += 1
                    
                    # 如果有多个连续相同的开发室，则合并
                    if i - start_index > 1:
                        # 合并开发室列的单元格
                        ws.merge_cells(
                            start_row=start_row + start_index,
                            start_column=7,
                            end_row=start_row + i - 1,
                            end_column=7
                        )
                        
                        # 设置合并后单元格的值和格式
                        cell = ws.cell(
                            row=start_row + start_index, 
                            column=7, 
                            value=current_room
                        )
                        cell.alignment = Alignment(
                            horizontal='center', 
                            vertical='center', 
                            wrap_text=True
                        )
                        cell.border = thin_border
                    else:
                        # 单个开发室，只设置格式
                        cell = ws.cell(
                            row=start_row + start_index, 
                            column=7, 
                            value=current_room
                        )
                        cell.alignment = Alignment(
                            horizontal='center', 
                            vertical='center', 
                            wrap_text=True
                        )
                        cell.border = thin_border
            
            # 设置固定的列宽
            ws.column_dimensions['A'].width = 10   # 序号列
            ws.column_dimensions['B'].width = 15   # 奖惩项点列
            ws.column_dimensions['C'].width = 25   # 奖惩明细列
            ws.column_dimensions['D'].width = 15   # 建议奖励金额列
            ws.column_dimensions['E'].width = 15   # 建议处罚金额列
            ws.column_dimensions['F'].width = 15   # 涉及人员列
            ws.column_dimensions['G'].width = 20   # 开发室列
            
            # 自动调整行高（基于内容）
            for row in ws.iter_rows():
                max_height = 15  # 默认行高
                for cell in row:
                    if cell.value:
                        # 计算行高，假设每行15磅，每换行增加15磅
                        lines = str(cell.value).count('\n') + 1
                        cell_height = lines * 15
                        if cell_height > max_height:
                            max_height = cell_height
                ws.row_dimensions[row[0].row].height = max_height
            
            # 6. 保存文件
            # 确定输出路径：如果指定了输出目录则使用，否则使用源文件所在目录
            output_path = os.path.join(
                self.output_dir if self.output_dir else os.path.dirname(self.source_file),
                "result.xlsx"
            )
            
            wb.save(output_path)
            if log_callback:
                log_callback(f"数据处理完成，结果已保存到: {output_path}")
            
            return output_path
            
        except Exception as e:
            # 捕获并重新抛出异常，添加更详细的错误信息
            raise Exception(f"处理过程中发生错误: {str(e)}")